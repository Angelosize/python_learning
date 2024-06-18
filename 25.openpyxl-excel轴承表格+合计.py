import openpyxl
lastweekzh=eval(input("上个月原有轴承多少件:"))
thisweekzh=eval(input("这个月新进轴承多少件:"))
getoutzh=eval(input("取走了轴承多少件:"))
lastweekdj=eval(input("上个月原有电机多少件:"))
thisweekdj=eval(input("这个月新进电机多少件:"))
getoutdj=eval(input("取走了电机多少件:"))
lst=[
    ['类别','原有库存件数','本月新进件数','本月支出件数','本月剩余件数'],
    ['轴承',lastweekzh,thisweekzh,getoutzh,lastweekzh+thisweekzh-getoutzh],
    ['电机',lastweekdj,thisweekdj,getoutdj,lastweekdj+thisweekdj-getoutdj],
    ['合计',lastweekzh+lastweekdj,thisweekzh+thisweekdj,getoutzh+getoutdj,lastweekzh+thisweekzh-getoutzh+lastweekdj+thisweekdj-getoutdj]
]
workbook=openpyxl.Workbook()
sheet=workbook.create_sheet('我爱轴承')
for it in lst:
    sheet.append(it)
print(workbook.sheetnames)
del workbook["Sheet"]
print(workbook.sheetnames)
workbook.save('python_files\我爱轴承.xlsx')