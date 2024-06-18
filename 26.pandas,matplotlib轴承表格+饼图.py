import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
lastweekzh=eval(input("上个月原有轴承多少件:"))
thisweekzh=eval(input("这个月新进轴承多少件:"))
getoutzh=eval(input("取走了轴承多少件:"))
lastweekdj=eval(input("上个月原有电机多少件:"))
thisweekdj=eval(input("这个月新进电机多少件:"))
getoutdj=eval(input("取走了电机多少件:"))
lst=[
    ['类别','原有库存件数','本月新进件数','本月支出件数','本月剩余件数'],
    ['轴承',lastweekzh,thisweekzh,getoutzh,lastweekzh+thisweekzh-getoutzh],
    ['电机',lastweekdj,thisweekdj,getoutdj,lastweekdj+thisweekdj-getoutdj]
]
workbook=openpyxl.Workbook()
sheet=workbook.create_sheet('我爱轴承')
for it in lst:
    sheet.append(it)
# print(workbook.sheetnames)
del workbook["Sheet"]
# print(workbook.sheetnames)
workbook.save('python_files\我爱轴承.xlsx')
df=pd.read_excel('python_files\我爱轴承.xlsx')
# print(df)
plt.rcParams['font.sans-serif']=['SimHei']
plt.figure(figsize=(10,6))
labels=df['类别']
y=df['原有库存件数']
# print(labels,'\n',y,sep='')
plt.pie(y,labels=labels,autopct='%1.1f%%',startangle=90)
plt.axis('equal')
plt.title('原有库存件数')
plt.show()


plt.rcParams['font.sans-serif']=['SimHei']
plt.figure(figsize=(10,6))
labels=df['类别']
y=df['本月新进件数']
# print(labels,'\n',y,sep='')
plt.pie(y,labels=labels,autopct='%1.1f%%',startangle=90)
plt.axis('equal')
plt.title('本月新进件数')
plt.show()

plt.rcParams['font.sans-serif']=['SimHei']
plt.figure(figsize=(10,6))
labels=df['类别']
y=df['本月支出件数']
# print(labels,'\n',y,sep='')
plt.pie(y,labels=labels,autopct='%1.1f%%',startangle=90)
plt.axis('equal')
plt.title('本月支出件数')
plt.show()


plt.rcParams['font.sans-serif']=['SimHei']
plt.figure(figsize=(10,6))
labels=df['类别']
y=df['本月剩余件数']
# print(labels,'\n',y,sep='')
plt.pie(y,labels=labels,autopct='%1.1f%%',startangle=90)
plt.axis('equal')
plt.title('本月剩余件数')
plt.show()
